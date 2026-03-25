"""
バッテリー交換実績集計アプリ
PT企業(user_company)毎に、user_nameと自転車メーカー別の集計を行います。
"""
import streamlit as st
import pandas as pd
import plotly.express as px
import io
import os
import zipfile



from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

from constants import TEMP_COLS, BIKE_COMPANY_EXCLUDE_COLS, SELF_EXCHANGE_MAPPING
from processing import aggregate_by_company_and_maker

# ページ設定
st.set_page_config(
    page_title="バッテリー交換実績集計",
    page_icon="🔋",
    layout="wide"
)

# 認証機能（パブリックデプロイ時は常に有効）
# ローカル開発時のみ無効化する場合は環境変数 DISABLE_AUTH=true を設定
ENABLE_AUTH = os.getenv("DISABLE_AUTH", "false").lower() != "true"

if ENABLE_AUTH:
    try:
        from auth import check_password, logout, get_authenticated_user

        if not check_password():
            st.stop()

        # ログアウトボタンをサイドバーに追加
        with st.sidebar:
            st.markdown("---")
            st.markdown(f"👤 ログイン中: **{get_authenticated_user()}**")
            if st.button("🚪 ログアウト"):
                logout()
    except Exception as e:
        st.error(f"🔐 認証モジュールのエラー: {e}")
        st.error("Secretsに [passwords] セクションが設定されているか確認してください")
        st.code("""
[passwords]
admin = "パスワードハッシュ値"
        """)
        st.stop()


def _apply_excel_table(ws, table_name: str, style: str = "TableStyleMedium9"):
    """ワークシートのデータ範囲をExcelテーブルとして設定する"""
    if ws.max_row < 2:
        return
    ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    tbl = Table(displayName=table_name, ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(
        name=style, showRowStripes=True,
        showFirstColumn=False, showLastColumn=False, showColumnStripes=False
    )
    ws.add_table(tbl)


def _add_pivot_sheet(writer, df_pivot: pd.DataFrame, e_col_name: str,
                     table_name: str = "PivotTable1", self_exc_df: pd.DataFrame = None):
    """bike_company × 自転車メーカー名 のクロス集計シートを追加する。
    self_exc_df が指定された場合は '交換種別' 行を追加して自社交換分も表示する。
    """
    maker_col = '自転車メーカー名'
    if e_col_name not in df_pivot.columns or maker_col not in df_pivot.columns or df_pivot.empty:
        return

    has_self = self_exc_df is not None and not self_exc_df.empty and maker_col in self_exc_df.columns

    if has_self:
        normal = df_pivot[[e_col_name, maker_col]].copy()
        normal['交換種別'] = '通常'
        self_part = self_exc_df[[e_col_name, maker_col]].copy()
        self_part['交換種別'] = '自社交換'
        combined = pd.concat([normal, self_part], ignore_index=True)
        pivot = pd.crosstab(
            [combined['交換種別'], combined[e_col_name]],
            combined[maker_col],
            margins=True,
            margins_name='合計'
        )
        pivot.index.names = ['交換種別', e_col_name]
    else:
        pivot = pd.crosstab(
            df_pivot[e_col_name],
            df_pivot[maker_col],
            margins=True,
            margins_name='合計'
        )
        pivot.index.name = e_col_name

    pivot = pivot.reset_index()
    pivot.to_excel(writer, sheet_name='ピボット', index=False)
    ws = writer.sheets['ピボット']
    _apply_excel_table(ws, table_name, style="TableStyleMedium2")


@st.cache_data(show_spinner=False)
def load_excel_from_uploaded_file(uploaded_file) -> pd.DataFrame:
    """アップロードされたファイルを読み込む"""
    try:
        df = pd.read_excel(uploaded_file)
        # E列（インデックス4）とV列（インデックス21）の列名を保存
        if len(df.columns) > 4:
            df.attrs['e_column_name'] = df.columns[4]  # E列
        if len(df.columns) > 21:
            df.attrs['v_column_name'] = df.columns[21]  # V列（user_company(所属)）
        return df
    except Exception as e:
        st.error(f"ファイル読み込みエラー: {e}")
        return None





def main():
    st.title("🔋 バッテリー交換実績集計アプリ")
    st.markdown("---")

    # サイドバー：ファイル選択
    with st.sidebar:
        st.header("⚙️ 設定")

        # ファイルアップロード
        uploaded_file = st.file_uploader(
            "Excelファイルをアップロード",
            type=['xlsx', 'xls'],
            help="バッテリー交換実績データのExcelファイルをアップロードしてください"
        )



        # バージョン情報（デバッグ用）
        st.markdown("---")
        st.caption("Version: 2026-03-25-v12 (bike_company選択ダウンロード対応)")

    # メインエリア
    if uploaded_file is not None:
        # アップロードされたファイルを読み込み
        with st.spinner("📂 ファイルを読み込み中..."):
            df = load_excel_from_uploaded_file(uploaded_file)

        if df is not None:
            st.success(f"✅ ファイル読み込み完了: {len(df):,}行のデータ")



            # データプレビュー
            with st.expander("📊 データプレビュー（最初の10行）"):
                st.dataframe(df.head(10))

            # 集計実行ボタン
            if st.button("🔄 集計実行", type="primary", use_container_width=True):
                progress_bar = st.progress(0, text="集計を開始しています...")
                status_text = st.empty()

                try:
                    status_text.text("📊 データを分析中...")
                    progress_bar.progress(10, text="重複データを検出中...")

                    status_text.text(f"🔍 重複データを検出中...（{len(df):,}行）")
                    progress_bar.progress(30, text="重複チェック実行中...")

                    # 集計実行（重複検出を含む、自社交換分を除外）
                    aggregated_data, aggregated_by_bike, self_exchange_df = aggregate_by_company_and_maker(df)

                    progress_bar.progress(90, text="集計結果を準備中...")
                    st.session_state['aggregated_data'] = aggregated_data
                    st.session_state['aggregated_by_bike'] = aggregated_by_bike
                    st.session_state['self_exchange_df'] = self_exchange_df

                    progress_bar.progress(100, text="完了！")
                    status_text.empty()
                    progress_bar.empty()

                    self_exchange_count = len(self_exchange_df) if not self_exchange_df.empty else 0
                    success_msg = f"✅ 集計完了！{len(aggregated_data)}社のデータを集計しました"
                    if self_exchange_count > 0:
                        success_msg += f"（自社交換分: {self_exchange_count:,}件を除外）"
                    st.success(success_msg)
                    st.balloons()

                except Exception as e:
                    progress_bar.empty()
                    status_text.empty()
                    st.error(f"❌ 集計エラー: {e}")
                    import traceback
                    with st.expander("詳細なエラー情報"):
                        st.code(traceback.format_exc())

            # 集計結果の表示
            if 'aggregated_data' in st.session_state:
                aggregated_data = st.session_state['aggregated_data']

                st.markdown("---")
                st.header("📈 集計結果")

                # 全企業のデータを1つのExcelファイルに出力
                st.subheader("全PT企業のデータを一括ダウンロード")

                st.info("💡 全PT企業の集計結果と生データを含むZIPファイルをダウンロードできます")
                st.warning("⚠️ 生データを含むため、ファイルサイズが大きくなります")

                if st.button("📦 全企業のExcelファイルを準備", key="prepare_all_excel"):
                    with st.spinner(f"全{len(aggregated_data)}社のExcelファイルをZIP化中..."):
                        zip_buffer = io.BytesIO()

                        # 生データから一時列を削除
                        df_clean = df.copy()
                        df_clean = df_clean.drop(columns=[col for col in TEMP_COLS if col in df_clean.columns], errors='ignore')

                        # 自社交換分のデータを準備
                        self_exchange_clean = None
                        if 'self_exchange_df' in st.session_state and not st.session_state['self_exchange_df'].empty:
                            self_exchange_clean = st.session_state['self_exchange_df'].copy()
                            self_exchange_clean = self_exchange_clean.drop(columns=[col for col in TEMP_COLS if col in self_exchange_clean.columns], errors='ignore')

                        today_str = pd.Timestamp.now().strftime('%Y%m%d')

                        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                            progress_bar = st.progress(0)
                            total = len(aggregated_data) + 1  # 企業ごとのファイル + 全企業まとめファイル

                            # 全企業の集計結果をまとめたExcelファイルを作成
                            all_companies_excel = io.BytesIO()
                            with pd.ExcelWriter(all_companies_excel, engine='openpyxl') as writer:
                                all_companies_data = []

                                for company, data in aggregated_data.items():
                                    company_with_name = data.copy()
                                    company_with_name.insert(0, 'PT企業名', company)
                                    all_companies_data.append(company_with_name)

                                combined_df = pd.concat(all_companies_data, ignore_index=True)
                                combined_df.to_excel(writer, sheet_name='全PT企業集計', index=False)
                                _apply_excel_table(writer.sheets['全PT企業集計'], "TableAllCompanies")

                                if self_exchange_clean is not None and not self_exchange_clean.empty:
                                    self_exchange_clean.to_excel(writer, sheet_name='自社交換分', index=False)
                                    _apply_excel_table(writer.sheets['自社交換分'], "TableSelfExchange")

                            all_companies_excel.seek(0)
                            zip_file.writestr(f"全企業_集計結果_{today_str}.xlsx", all_companies_excel.getvalue())
                            progress_bar.progress(1 / total)

                            # 各企業ごとに1つのExcelファイルを作成
                            for idx, (company, data) in enumerate(aggregated_data.items()):
                                excel_buffer = io.BytesIO()

                                with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                                    data.to_excel(writer, sheet_name='集計結果', index=False)
                                    _apply_excel_table(writer.sheets['集計結果'], f"TableSummary{idx}")

                                    company_raw = df_clean[df_clean['user_company(所属)'] == company].copy()
                                    company_raw.to_excel(writer, sheet_name='生データ', index=False)
                                    _apply_excel_table(writer.sheets['生データ'], f"TableRaw{idx}")

                                    if self_exchange_clean is not None and not self_exchange_clean.empty:
                                        company_self_exchange = self_exchange_clean[self_exchange_clean['user_company(所属)'] == company].copy()
                                        if not company_self_exchange.empty:
                                            company_self_exchange.to_excel(writer, sheet_name='自社交換分', index=False)
                                            _apply_excel_table(writer.sheets['自社交換分'], f"TableSelf{idx}")

                                safe_company_name = company.replace('/', '_').replace('\\', '_')
                                zip_file.writestr(
                                    f"{safe_company_name}_集計結果_{today_str}.xlsx",
                                    excel_buffer.getvalue()
                                )

                                progress_bar.progress((idx + 2) / total)

                            progress_bar.empty()

                        zip_buffer.seek(0)
                        st.session_state['all_excel_data'] = zip_buffer.getvalue()
                        st.session_state['all_excel_filename'] = "全PT企業_集計結果_生データ.zip"
                        st.session_state['all_excel_mime'] = "application/zip"
                        st.success(f"✅ ZIPファイルの準備完了！（全企業_集計結果.xlsx + {len(aggregated_data)}社のExcelファイル）")

                # ダウンロードボタンを表示
                if 'all_excel_data' in st.session_state:
                    st.download_button(
                        label=f"📥 {st.session_state['all_excel_filename']} をダウンロード",
                        data=st.session_state['all_excel_data'],
                        file_name=st.session_state['all_excel_filename'],
                        mime=st.session_state.get('all_excel_mime', 'application/zip'),
                        key="download_all_excel"
                    )

                st.markdown("---")
                st.subheader("🚲 bike_company毎のローデータをダウンロード")
                st.info("💡 ダウンロードしたいbike_companyを選択してExcelを生成できます")

                e_col_name = df.attrs.get('e_column_name', None) if hasattr(df, 'attrs') else None
                if not e_col_name and len(df.columns) > 4:
                    e_col_name = df.columns[4]

                if e_col_name and e_col_name in df.columns:
                    df_clean_bike = df.copy()
                    exclude_cols = [col for col in TEMP_COLS + BIKE_COMPANY_EXCLUDE_COLS if col in df_clean_bike.columns]
                    df_clean_bike = df_clean_bike.drop(columns=exclude_cols, errors='ignore')

                    bike_companies = sorted(df_clean_bike[e_col_name].dropna().unique().tolist())

                    selected_bike_companies = st.multiselect(
                        "ダウンロードするbike_companyを選択",
                        options=bike_companies,
                        default=[],
                        placeholder="会社を選択してください（複数選択可）"
                    )

                    btn_label = f"📊 選択した{len(selected_bike_companies)}社のExcelを準備" if selected_bike_companies else "📊 Excelを準備（会社を選択してください）"
                    if st.button(btn_label, key="prepare_bike_excels", disabled=not selected_bike_companies):
                        with st.spinner(f"{len(selected_bike_companies)}社分のExcelファイルを生成中..."):
                            bike_excel_files = {}
                            for idx, bike_company in enumerate(selected_bike_companies):
                                excel_buf = io.BytesIO()
                                with pd.ExcelWriter(excel_buf, engine='openpyxl') as writer:
                                    if bike_company in SELF_EXCHANGE_MAPPING:
                                        # session_stateの自社交換分データを使って分離
                                        self_exchange_all = st.session_state.get('self_exchange_df', pd.DataFrame())
                                        if not self_exchange_all.empty and e_col_name in self_exchange_all.columns:
                                            self_exc_clean = self_exchange_all[
                                                self_exchange_all[e_col_name] == bike_company
                                            ].copy()
                                            self_exc_clean = self_exc_clean.drop(
                                                columns=[c for c in TEMP_COLS + BIKE_COMPANY_EXCLUDE_COLS if c in self_exc_clean.columns],
                                                errors='ignore'
                                            )
                                            # 非自社交換分 = 全行からインデックスで自社交換行を除外
                                            all_company = df_clean_bike[df_clean_bike[e_col_name] == bike_company]
                                            non_self_clean = all_company.drop(index=self_exc_clean.index, errors='ignore')
                                        else:
                                            self_exc_clean = pd.DataFrame()
                                            non_self_clean = df_clean_bike[df_clean_bike[e_col_name] == bike_company].copy()
                                        non_self_clean.to_excel(writer, sheet_name='ローデータ', index=False)
                                        _apply_excel_table(writer.sheets['ローデータ'], f"TableRaw{idx}")
                                        if not self_exc_clean.empty:
                                            self_exc_clean.to_excel(writer, sheet_name='自社交換分', index=False)
                                            _apply_excel_table(writer.sheets['自社交換分'], f"TableSelf{idx}")
                                        _add_pivot_sheet(writer, non_self_clean, e_col_name, f"PivotTable{idx}",
                                                         self_exc_df=self_exc_clean if not self_exc_clean.empty else None)
                                    else:
                                        company_raw = df_clean_bike[df_clean_bike[e_col_name] == bike_company].copy()
                                        company_raw.to_excel(writer, sheet_name='ローデータ', index=False)
                                        _apply_excel_table(writer.sheets['ローデータ'], f"TableRaw{idx}")
                                        _add_pivot_sheet(writer, company_raw, e_col_name, f"PivotTable{idx}")
                                excel_buf.seek(0)
                                bike_excel_files[str(bike_company)] = excel_buf.getvalue()
                            st.session_state['bike_excel_files'] = bike_excel_files
                            st.session_state['bike_excel_selected'] = list(selected_bike_companies)
                        st.success(f"✅ {len(selected_bike_companies)}社分のExcelファイルを準備しました")

                    if 'bike_excel_files' in st.session_state and st.session_state['bike_excel_files']:
                        st.caption(f"準備済み: {', '.join(st.session_state.get('bike_excel_selected', []))}")
                        cols = st.columns(3)
                        for i, bike_company in enumerate(sorted(st.session_state['bike_excel_files'].keys())):
                            safe_name = bike_company.replace('/', '_').replace('\\', '_')
                            with cols[i % 3]:
                                st.download_button(
                                    label=f"📥 {bike_company}",
                                    data=st.session_state['bike_excel_files'][bike_company],
                                    file_name=f"{safe_name}_ローデータ.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    key=f"download_bike_{i}"
                                )
                else:
                    st.warning("⚠️ E列（bike_company列）が見つかりません")

    else:
        # ファイルが選択されていない場合
        st.info("👈 サイドバーからExcelファイルをアップロードしてください")


if __name__ == "__main__":
    main()
